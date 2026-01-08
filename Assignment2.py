from openpyxl import load_workbook
import math
import time

class Airport:
    def __init__(self, id,latitude,longitude,runway, time_demand):
        self.id =id
        self.lat = latitude
        self.lon = longitude
        self.Run = runway
        self.Dt = time_demand

class Arc:
    def __init__(self,id, origin, destination, demand, distance):
        self.id = id
        self.From = origin
        self.To = destination
        self.Dem = demand
        self.Dist = distance

class Aircraft :
    def __init__(self,type, speed, seats, TAT, max_range, runway, lease_cost, fixed_op_cost, hour_cost, fuel_price, fleet):
        self.type = type
        self.V = speed
        self.s = seats
        self.TAT = TAT
        self.Range = max_range
        self.Runway = runway
        self.C_L = lease_cost
        self.C_op = fixed_op_cost
        self.C_h = hour_cost
        self.f = fuel_price
        self.N = fleet

class Node:
    def __init__(self,id,minute,destination,time_arrival,BT_totali,pax,state):
        self.id = id
        self.t = minute
        self.To = destination
        self.state = state
        self.tf = time_arrival
        self.pax = pax
        self.BT = BT_totali


def arc_define(airports):
    arc_demand = read_file_demand('DemandGroup11.xlsx','Group 11')
    arcs = {}
    List_arcs = []
    for ac in arc_demand:
        List_arcs.append(ac)
    i=0
    for ac in List_arcs:
        dist = calculate_distance(ac,airports)
        ac = ac + (dist,) 
        List_arcs[i] = ac
        i+=1       
    for (id, origin, destination, demand, distance)  in List_arcs:
        for ap in airports:
            if origin == ap.id:
                demand = [demand*x for x in ap.Dt ]
        ac = Arc(id, origin, destination, demand, distance)
        arcs[(origin,destination)] = ac

    
    return arcs

def aircraft_define():
    aircrafts = []
    type_info = read_file_aircraft('FleetType.xlsx','Fleet Type')
    for (type, speed, seats, TAT, max_range, runway, lease_cost, op_cost, hour_cost, fuel_price, fleet) in type_info:
        ac = Aircraft(type, speed, seats, TAT, max_range, runway, lease_cost, op_cost, hour_cost, fuel_price, fleet)
        aircrafts.append(ac)
    return aircrafts

def airport_define():
    airports = []
    arc_coordinates = read_file_coordinates('DemandGroup11.xlsx','Group 11')
    Dict_Dt = read_file_Dt('HourCoefficients.xlsx','Hour Coefficients')
    for (id,latitude,longitude,runway) in arc_coordinates:
        Dt = Dict_Dt[id]
        ap = Airport(id,latitude,longitude,runway,Dt)
        airports.append(ap)
    return airports

def dynamic_programming_iteration (arcs,airports,aircrafts):
    hub = 'EHAM'
    LF = 0.8
    #create dictionary to avoid for loop
    arc_pair = {}
    #
    nodes_aircraft = {}
    for ac in aircrafts:
        if ac.N > 0:
            nodes = []
            nodes.append(Node(hub,240,'End-Day',None,0,0,0)) # aircrafts should stay at the hub at the end of the day
            for t in range(239,-1,-1):
                h = math.floor(t/10)
                dest_nodes = nodes.copy()
                for ap in airports:
                    possible_state = []
                    for nd in dest_nodes:
                        if nd.id == ap.id:
                            if nd.t == t+1:
                                arc_profit = nd.state
                                node = Node(ap.id,t,nd.id,t+1,nd.BT,0,arc_profit)
                                possible_state.append(node)
                        else:
                            if ap.id == hub or nd.id == hub:
                                arc = arcs[(ap.id,nd.id)]
                                BT = (arc.Dist/ac.V*60+30)/6
                                TAT = ac.TAT/6
                                tf = math.ceil(BT+TAT) + t
                                # Check range constraint AND runway constraint
                                origin_runway = next((a.Run for a in airports if a.id == ap.id), 0)
                                dest_runway = next((a.Run for a in airports if a.id == nd.id), 0)
                                min_runway = min(origin_runway, dest_runway)
                                if tf == nd.t and ac.Range >= arc.Dist and ac.Runway <= min_runway:
                                    demand = arc.Dem[h]
                                    if h-1>0:
                                        demand += arc.Dem[h-1]
                                    if h-2 > 0:
                                        demand += arc.Dem[h-2]
                                    demand = math.floor(demand)
                                    seats_av = int(math.floor(ac.s*LF))
                                    if seats_av >= demand:
                                        pax = demand
                                    else:
                                        pax = seats_av
                                    Y = 5.9*arc.Dist**(-0.76)+0.043
                                    revenue = pax*arc.Dist*Y
                                    fuel_price = 1.42  # USD/gallon
                                    cost = ac.C_op + ac.C_h*arc.Dist/ac.V + (ac.f*fuel_price*arc.Dist)/1.5 
                                    arc_profit = revenue-cost+nd.state
                                    node = Node(ap.id,t,nd.id,tf,BT+nd.BT,pax,arc_profit)
                                    possible_state.append(node)
                    
                    if possible_state != []:
                        max_state = max(possible_state, key=lambda n: n.state)
                        max_state.state = max_state.state
                        nodes.append(max_state)
            nodes_aircraft[ac.type] = nodes
    return nodes_aircraft

def find_best_route (nodes_iteration,aircrafts):
    hub = 'EHAM'
    possible_routes = {}
    for ac in aircrafts:
            if ac.N >0:
                ac_type = ac.type
                nodes = {}
                route = []
                for i in range(240,-1,-1):
                    nodes[i] = []
                for nd in nodes_iteration[ac_type]:
                    if nd.t == 0:
                        nd.state -= ac.C_L
                    nodes[nd.t].append(nd)
                # Filter to only nodes starting at the hub
                hub_nodes = [nd for nd in nodes[0] if nd.id == hub]
                if not hub_nodes:
                    continue  # No valid routes for this aircraft type
                final_point = max(hub_nodes, key = lambda nd: nd.state)
                while final_point.BT <= 60 and len(hub_nodes) > 1:
                    hub_nodes.remove(final_point)
                    final_point = max(hub_nodes, key = lambda nd: nd.state)
                route.append(final_point)
                prev_nd = final_point
                while prev_nd.t < 240:
                    next_nd = prev_nd.To
                    for nd in nodes[prev_nd.tf]:
                        if nd.id == next_nd:
                            route.append(nd)
                            prev_nd = nd
                            break
                possible_routes[ac.type] = route
    
    best_type, best_route = max(possible_routes.items(), key=lambda kv: max((nd.state for nd in kv[1] if nd.t == 0), default=-float("inf")))
    best_state_t0 = max(nd.state for nd in best_route if nd.t == 0)


    return best_type,best_route,best_state_t0

def routes_defnition (arcs,airports,aircrafts):
    total_fleet = 0
    for ac in aircrafts:
        total_fleet += ac.N
    routes = {}
    i = 1
    while total_fleet > 0:
        start_it = time.time()
        iteration = dynamic_programming_iteration(arcs,airports,aircrafts)
        end_it = time.time()
        delta_it = end_it - start_it
        best_type,best_route,best_state_t0 = find_best_route (iteration,aircrafts)
        
        # Only add route if it's profitable
        if best_state_t0 <= 0:
            print(f"iteration {i}: No more profitable routes found. Stopping.")
            break
            
        print(f"iteration {i} = {delta_it:.2f}s | {best_type} | Profit: €{best_state_t0:,.2f}")
        for ac in aircrafts:
            if ac.type == best_type:
                ac.N += -1
        routes[(best_type,i)] = (best_route,best_state_t0)
        for nd in best_route:
            if nd.id != nd.To:
                h = math.floor(nd.t/10)
                arc = arcs.get((nd.id,nd.To))
                if arc != None:
                    if arc.Dem[h]>nd.pax:
                        arc.Dem[h] += -(nd.pax)
                    else: 
                        if arc.Dem[h-1] > nd.pax-arc.Dem[h]:
                            arc.Dem[h-1] -= nd.pax-arc.Dem[h]
                            arc.Dem[h] = 0
                        else:
                            arc.Dem[h-1] = 0
                            if arc.Dem[h-2] > nd.pax-arc.Dem[h]-arc.Dem[h-1]:
                                arc.Dem[h-2] -= nd.pax-arc.Dem[h]-arc.Dem[h-1]
                                arc.Dem[h] = 0
                                arc.Dem[h-1] = 0
                            else:
                                arc.Dem[h] = 0
                                arc.Dem[h-1] = 0
                                arc.Dem[h-2] = 0
        i += 1
        total_fleet-=1
    return routes



#########################################################################

#########################################################################
def read_file_demand(file_name, sheet_name):  #read a matrix to combine the values for each arc
    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]


    header = [c for c in next(ws.iter_rows(min_row=11,max_row = 11,values_only=True))]
    airports = []
    for v in header[2 :]:
        if v is None:
            break
        airports.append(v)  
    
    data = []
    id = 0
    for row in ws.iter_rows(min_row=12, values_only=True):
        origin = row[1]
        values = row[2:]

        for dest, val in zip(airports, values):
            if origin == dest:
                continue
            if val is None or val == 0:
                continue
            data.append((id,origin,dest,val))
            id +=1

    return data


def read_file_aircraft(file_name,sheet_name): #read a file to obtain the value for aircraft specification

    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]
    List = []
    for cols in ws.iter_cols(min_col=2, max_col = 4, values_only = True):
        values = cols[0:]
        List.append((values))
    return List

def read_file_Dt(file_name, sheet_name):
    
    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]
    Dict = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        airport_id = row[2]
        values = list(row[3:])  
        Dict[airport_id] = values
    return Dict

def read_file_coordinates(file_name,sheet_name):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]
    List = []
    for cols in ws.iter_cols(min_col=3, values_only = True):
        values = cols[4:8]
        List.append((values))
    return List

def calculate_distance (a,airport):
    R_e = 6371 #[km]
    k = 0
    for ap in airport:
            if a[1] == ap.id:
                k+=1
                lam_from = math.radians(ap.lon)
                phi_from = math.radians(ap.lat)
            elif a[2] == ap.id:
                k+=1
                lam_to = math.radians(ap.lon)
                phi_to = math.radians(ap.lat)
            if k == 2:
                break
        
    delta = 2*math.asin(math.sqrt(math.sin((phi_from-phi_to)/2)**2+math.cos(phi_from)*math.cos(phi_to)*(math.sin((lam_from-lam_to)/2))**2))
    dist_a = delta*R_e
    return (dist_a)


def print_results_summary(routes, arcs_original, airports):
    """Print a comprehensive summary of the aircraft routing solution."""
    print("\n" + "="*80)
    print("                    AIRCRAFT ROUTING SOLUTION SUMMARY")
    print("="*80)
    
    total_profit = 0
    total_revenue = 0
    total_cost = 0
    total_pax = 0
    total_flights = 0
    aircraft_summary = {}
    
    for (ac_type, route_num), (route, profit) in routes.items():
        if ac_type not in aircraft_summary:
            aircraft_summary[ac_type] = {'routes': [], 'total_profit': 0, 'total_flights': 0, 'total_pax': 0}
        aircraft_summary[ac_type]['routes'].append((route_num, route, profit))
        aircraft_summary[ac_type]['total_profit'] += profit
    
    print(f"\n{'Aircraft Type':<25} {'Routes Used':<15} {'Total Profit (€)':<20}")
    print("-"*60)
    for ac_type, data in aircraft_summary.items():
        print(f"{ac_type:<25} {len(data['routes']):<15} {data['total_profit']:>15,.2f}")
        total_profit += data['total_profit']
    
    print("\n" + "="*80)
    print("                         DETAILED FLIGHT SCHEDULE")
    print("="*80)
    
    for (ac_type, route_num), (route, profit) in sorted(routes.items()):
        print(f"\n--- Aircraft: {ac_type} | Route #{route_num} | Daily Profit: €{profit:,.2f} ---")
        print(f"{'Dep Time':<12} {'Origin':<8} {'Dest':<8} {'Arr Time':<12} {'Pax':<6} {'Flight Time (min)'}")
        print("-"*70)
        
        route_flights = 0
        route_pax = 0
        route_block_time = 0
        for nd in route:
            if nd.id != nd.To and nd.To != 'End-Day':
                dep_hour = (nd.t * 6) // 60
                dep_min = (nd.t * 6) % 60
                arr_hour = (nd.tf * 6) // 60 if nd.tf else 0
                arr_min = (nd.tf * 6) % 60 if nd.tf else 0
                flight_time = (nd.tf - nd.t) * 6 if nd.tf else 0  # Individual flight time
                
                print(f"{dep_hour:02d}:{dep_min:02d}        {nd.id:<8} {nd.To:<8} {arr_hour:02d}:{arr_min:02d}        {nd.pax:<6} {flight_time:.0f}")
                route_flights += 1
                route_pax += nd.pax
                route_block_time += flight_time
                total_pax += nd.pax
        
        total_flights += route_flights
        print(f"  Subtotal: {route_flights} flights, {route_pax} passengers, {route_block_time:.0f} min total block time")
    
    print("\n" + "="*80)
    print("                           OVERALL STATISTICS")
    print("="*80)
    print(f"  Total Aircraft Used:      {len(routes)}")
    print(f"  Total Flights:            {total_flights}")
    print(f"  Total Passengers:         {total_pax}")
    print(f"  Total Daily Profit:       €{total_profit:,.2f}")
    print("="*80 + "\n")
    
    return total_profit


# ======== MAIN EXECUTION ========
if __name__ == "__main__":
    airports = airport_define()
    arcs = arc_define(airports)
    arcs_original = {k: Arc(v.id, v.From, v.To, v.Dem.copy(), v.Dist) for k, v in arcs.items()}  # Keep copy for summary
    aircrafts = aircraft_define()
    
    print('='*60)
    print('   AIRCRAFT ROUTING OPTIMIZATION - DYNAMIC PROGRAMMING')
    print('='*60)
    print(f'Hub Airport: EHAM (Amsterdam)')
    print(f'Number of airports: {len(airports)}')
    print(f'Number of arcs: {len(arcs)}')
    print(f'Aircraft types available: {[ac.type for ac in aircrafts if ac.N > 0]}')
    print('='*60)
    print('\nStarting optimization...\n')
    
    start = time.time()
    routes = routes_defnition(arcs, airports, aircrafts)
    end = time.time()
    
    print(f"\nOptimization completed in {end-start:.2f} seconds")
    
    # Print the results summary
    print_results_summary(routes, arcs_original, airports)